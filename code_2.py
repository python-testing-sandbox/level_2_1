import configparser
import datetime
import os
import sqlite3
import sys
import re
from pathlib import Path
from typing import Optional, Set, Callable, Union, Collection, Mapping, Any

import deal as deal
from PIL import UnidentifiedImageError
from chardet import detect
from pytz import timezone as check_timezone
import pytz.exceptions
import xlrd
from openpyxl import Workbook, load_workbook

from level1_basics.code import get_image_height_in_pixels, flat

OBSCENE_BASE_TABLE_NAME = 'words'
CONFIG_SECTION_NAME = 'testing_sandbox'


class ColumnError(Exception):
    pass


def load_obscene_words(db_path: str) -> Set[str]:
    connection = sqlite3.connect(db_path)
    cursor = connection.cursor()
    return set(flat(cursor.execute(
        f'SELECT word FROM {OBSCENE_BASE_TABLE_NAME}',
    ).fetchall()))


def get_all_filepathes_recursively(path: str, extension: str) -> list[str]:
    pathlist = Path(path).glob(f'**/*.{extension}')
    return [
        str(p) for p in pathlist
        if not os.path.isdir(str(p))
    ]


def get_content_from_file(filepath: str, guess_encoding: bool) -> Optional[str]:
    if guess_encoding:
        with open(filepath, 'rb') as file_handler:
            binary_content = file_handler.read()
        encoding = detect(binary_content)['encoding']
        with open(filepath, 'r', encoding=encoding) as file_handler:
            return file_handler.read()
    else:
        try:
            with open(filepath, 'r') as file_handler:
                return file_handler.read()
        except UnicodeDecodeError:
            return None


def get_params_from_config(config_path: str) -> Mapping[str, Any]:
    config = configparser.ConfigParser()
    config.read(config_path)
    if not config.has_section(CONFIG_SECTION_NAME):
        return {}
    params = dict(config[CONFIG_SECTION_NAME])
    if 'processes' in params:
        params['processes'] = int(params['processes'])  # type: ignore
    if 'exclude' in params:
        params['exclude'] = params['exclude'].split(',')  # type: ignore
    if 'exit_zero' in params:
        params['exit_zero'] = params['exit_zero'] == 'True'  # type: ignore
    if 'reorder_vocabulary' in params:
        params['reorder_vocabulary'] = params['reorder_vocabulary'] == 'True'  # type: ignore
    if 'process_dots' in params:
        params['process_dots'] = params['process_dots'] == 'True'  # type: ignore
    if 'verbosity' in params:
        params['verbosity'] = int(params['verbosity'])  # type: ignore
    return params


def reorder_vocabulary(vocabulary_path: str) -> None:
    with open(vocabulary_path, 'r') as file_handler:
        raw_lines = file_handler.readlines()
    sections: list[list[str]] = []
    current_section: list[str] = []
    for line in raw_lines:
        processed_line = line.strip()
        if not processed_line:
            continue
        if processed_line.startswith('#') and current_section:
            sections.append(current_section)
            current_section = []
        current_section.append(processed_line)
    if current_section:
        sections.append(current_section)
    sorted_sections: list[list[str]] = []
    for section_num, section in enumerate(sections, 1):
        sorted_sections.append(
            [f'{r}\n' for r in section if r.startswith('#')]
            + sorted(f'{r}\n' for r in section if not r.startswith('#'))
            + (['\n'] if section_num < len(sections) else []),
        )

    with open(vocabulary_path, 'w') as file_handler:
        file_handler.writelines(flat(sorted_sections))


@deal.post(lambda r: all(u.startswith('http') for u in r))
def fetch_badges_urls(readme_content: str):
    if not readme_content:
        return []
    image_urls = re.findall(r'(?:!\[.*?\]\((.*?)\))', readme_content)
    max_badge_height = 60
    badges_urls = []
    for url in image_urls:
        try:
            height = get_image_height_in_pixels(url)
        except UnidentifiedImageError:  # this happens with svg, should parse it and get height
            badges_urls.append(url)
            continue
        if height and height < max_badge_height:
            badges_urls.append(url)
    return badges_urls


def fetch_detailed_pull_requests(api, open_pull_requests):
    pull_requests = {}
    for pull_request in open_pull_requests:
        pr = api.fetch_pull_request(pull_request['number'])
        if pr:
            pull_requests[pr['number']] = pr
    return pull_requests


class DateTimeProcessor:
    def __init__(self, formats: Collection[str] = None,
                 parser: Callable = None, timezone: Union[str, None] = None, **kwargs) -> None:
        super().__init__(**kwargs)
        self.formats = formats
        self.parser = parser
        self.user_timezone = None
        if timezone:
            try:
                self.user_timezone = check_timezone(timezone)
            except pytz.exceptions.UnknownTimeZoneError:
                self.user_timezone = None
                raise ValueError(f'Unknown time zone.')

    def process_value(self, value):
        if isinstance(value, str):
            value = self._get_datetime_from_string(value.strip())
        elif isinstance(value, datetime.datetime):
            pass
        elif isinstance(value, datetime.date):
            value = datetime.datetime.combine(value, datetime.time.min)
        else:
            raise ColumnError(f'Unable to convert to date {value}.')
        if self.user_timezone:
            print(self.user_timezone)
            value = value.astimezone(self.user_timezone)
        return value

    def _get_datetime_from_string(self, value: str) -> datetime.datetime:
        if not self.formats:
            try:
                return self.parser(value)
            except ValueError:
                raise ColumnError(f'Unable to convert "{value}" to date.')
        for date_format in self.formats:
            try:
                return datetime.datetime.strptime(value, date_format)
            except ValueError:
                pass
        raise ColumnError(f'Value "{value}" is not accordance with the format {self.formats}.')


def _load_workbook_from_xls(file_path, file_contents) -> Workbook:
    xls_workbook = xlrd.open_workbook(filename=file_path, file_contents=file_contents)
    xls_sheet = xls_workbook.sheet_by_index(0)
    nrows = xls_sheet.nrows
    ncols = xls_sheet.ncols
    wb = Workbook()
    ws = wb[wb.sheetnames[0]]

    for row in range(nrows):
        for col in range(ncols):
            cell = xls_sheet.cell(row, col)
            value = cell.value
            if value and cell.ctype == 3:
                value = datetime.datetime(*xlrd.xldate_as_tuple(value, xls_workbook.datemode))
                print(value)
            ws.cell(row=row + 1, column=col + 1).value = value

    return wb


def skip_exceptions_to_reraise():
    """Return a tuple of exceptions meaning 'skip this test', to re-raise.
    This is intended to cover most common test runners; if you would
    like another to be added please open an issue or pull request adding
    it to this function and to tests/cover/test_lazy_import.py
    """
    # This is a set because nose may simply re-export unittest.SkipTest
    exceptions = set()
    # We use this sys.modules trick to avoid importing libraries -
    # you can't be an instance of a type from an unimported module!
    # This is fast enough that we don't need to cache the result,
    # and more importantly it avoids possible side-effects :-)
    if "unittest" in sys.modules:
        exceptions.add(sys.modules["unittest"].SkipTest)
    if "unittest2" in sys.modules:  # pragma: no cover
        exceptions.add(sys.modules["unittest2"].SkipTest)
    if "nose" in sys.modules:  # pragma: no cover
        exceptions.add(sys.modules["nose"].SkipTest)
    if "_pytest" in sys.modules:  # pragma: no branch
        exceptions.add(sys.modules["_pytest"].outcomes.Skipped)
    return tuple(sorted(exceptions, key=str))


def _set_listed_at(item, marketplace) -> None:
    listed_at_field_name = f'{marketplace.value}_listed_at'
    if hasattr(item, listed_at_field_name):
        setattr(item, listed_at_field_name, datetime.datetime.now())
